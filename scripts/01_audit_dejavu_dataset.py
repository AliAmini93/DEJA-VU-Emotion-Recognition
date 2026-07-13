#!/usr/bin/env python
"""Read-only structural audit of the extracted DEJA-VU dataset.

Inspects directory layout, the SQLite database schema/rows, the XLSX
workbook's sheets/headers, HDF5 group/dataset/attribute structure (metadata
only — never loads full arrays), and raw XDF stream metadata (pyxdf must
parse the whole file internally, but this script discards the loaded arrays
immediately after extracting stream headers and does not retain or write out
any signal sample data).

Never modifies any input file. Performs no preprocessing, no segmentation,
no model training.

Writes:
  docs/data_audit_dejavu.json
  docs/dejavu_file_inventory.csv
  docs/dejavu_subject_session_summary.csv
  docs/dejavu_channel_inventory.csv
  docs/dejavu_event_inventory.csv
"""
from __future__ import annotations

import csv
import json
import sqlite3
import sys
from pathlib import Path

import h5py
import openpyxl
import pyxdf

sys.path.insert(0, str(Path(__file__).resolve().parent))
from dejavu_lib import parse_participant_session_from_path  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent
DATASET_ROOT = Path("/mnt/HDD/AliWorks/DEJA-VU/extracted/dataset/DEJA-VU")
DOCS_DIR = REPO_ROOT / "docs"

TRUNCATION_TOLERANCE = 0  # files must match archive-listed size exactly to be trusted


def categorize(rel: str) -> str:
    if rel.endswith(".xdf"):
        return "raw_xdf"
    if "/preprocessed/" in rel and rel.endswith(".h5"):
        return "preprocessed_hdf5"
    if "/segments/" in rel and rel.endswith(".h5"):
        return "segment_file"
    if rel.endswith(".db"):
        return "sqlite_database"
    if rel.endswith(".xlsx"):
        return "spreadsheet"
    if rel.endswith(".py") or rel.endswith(".yml") or rel.endswith(".txt"):
        return "official_code"
    if rel.endswith(".docx") or rel.endswith(".md"):
        return "documentation"
    return "other"


def load_truncation_set() -> set[str]:
    """Read the extraction inventory to know which files are truncated and
    must not be opened for content inspection."""
    truncated = set()
    inv_path = DOCS_DIR / "dejavu_extracted_file_inventory.csv"
    if not inv_path.exists():
        return truncated
    with open(inv_path, newline="") as f:
        for row in csv.DictReader(f):
            if row["size_match"] == "False":
                truncated.add(row["relative_path"])
    return truncated


def build_file_inventory(truncated: set[str]) -> list[dict]:
    rows = []
    for p in sorted(DATASET_ROOT.rglob("*")):
        if p.is_dir():
            continue
        rel = "DEJA-VU/" + str(p.relative_to(DATASET_ROOT))
        ident = parse_participant_session_from_path(str(p))
        rows.append({
            "relative_path": rel,
            "filename": p.name,
            "extension": p.suffix,
            "size_bytes": p.stat().st_size,
            "category": categorize(rel),
            "participant": ident["subject"] or "",
            "session": ident["session"] or "",
            "trusted": rel not in truncated,
        })
    return rows


def audit_sqlite(db_path: Path) -> dict:
    uri = f"file:{db_path}?mode=ro"
    conn = sqlite3.connect(uri, uri=True)
    try:
        cur = conn.cursor()
        cur.execute("SELECT name, type FROM sqlite_master WHERE type IN ('table','view') ORDER BY name")
        tables = [r[0] for r in cur.fetchall()]

        schema = {}
        for t in tables:
            cur.execute(f"PRAGMA table_info('{t}')")
            cols = [{"name": r[1], "type": r[2], "notnull": bool(r[3]), "pk": bool(r[5])} for r in cur.fetchall()]
            cur.execute(f"SELECT COUNT(*) FROM '{t}'")
            row_count = cur.fetchone()[0]
            schema[t] = {"columns": cols, "row_count": row_count}

        distinct_subjects = []
        distinct_sessions = []
        distinct_subject_session_pairs = 0
        if "videos" in tables:
            cur.execute("SELECT DISTINCT subject FROM videos ORDER BY subject")
            distinct_subjects = [r[0] for r in cur.fetchall()]
            cur.execute("SELECT DISTINCT session FROM videos ORDER BY session")
            distinct_sessions = [r[0] for r in cur.fetchall()]
            cur.execute("SELECT COUNT(*) FROM (SELECT DISTINCT subject, session FROM videos)")
            distinct_subject_session_pairs = cur.fetchone()[0]

        return {
            "tables": schema,
            "distinct_subjects": distinct_subjects,
            "distinct_subject_count": len(distinct_subjects),
            "distinct_sessions_seen": distinct_sessions,
            "distinct_subject_session_pairs": distinct_subject_session_pairs,
        }
    finally:
        conn.close()


def audit_xlsx(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheets = {}
        for name in wb.sheetnames:
            ws = wb[name]
            header = None
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                header = list(row)
                break
            sheets[name] = {
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "header_row": header,
            }
        return {"sheet_names": wb.sheetnames, "sheets": sheets}
    finally:
        wb.close()


def audit_hdf5_file(path: Path) -> dict:
    """Metadata-only: shapes/dtypes/attrs, never full array values."""
    result = {"groups": {}, "top_level_attrs": {}}
    with h5py.File(path, "r") as f:
        for k, v in f.attrs.items():
            result["top_level_attrs"][k] = _jsonable(v)
        for group_name in f.keys():
            grp = f[group_name]
            if isinstance(grp, h5py.Group):
                entry = {"attrs": {k: _jsonable(v) for k, v in grp.attrs.items()}, "datasets": {}}
                for ds_name in grp.keys():
                    ds = grp[ds_name]
                    if isinstance(ds, h5py.Dataset):
                        entry["datasets"][ds_name] = {
                            "shape": list(ds.shape),
                            "dtype": str(ds.dtype),
                            "compression": ds.compression,
                        }
                result["groups"][group_name] = entry
    return result


def _jsonable(v):
    if hasattr(v, "item"):
        try:
            return v.item()
        except Exception:
            pass
    if hasattr(v, "tolist"):
        return v.tolist()
    return v


def audit_xdf_file(path: Path) -> dict:
    """Parses the XDF file via pyxdf (unavoidable — pyxdf has no
    header-only mode) but discards loaded sample arrays immediately after
    extracting per-stream metadata."""
    streams, header = pyxdf.load_xdf(str(path), verbose=False)
    stream_summaries = []
    for s in streams:
        info = s["info"]
        name = info.get("name", ["Unknown"])[0]
        stype = info.get("type", ["Unknown"])[0]
        srate = float(info.get("nominal_srate", ["0"])[0])
        channel_count = int(info.get("channel_count", ["0"])[0])
        channel_labels = []
        try:
            channels = info["desc"][0]["channels"][0]["channel"]
            channel_labels = [ch["label"][0] for ch in channels]
        except (KeyError, IndexError, TypeError):
            pass
        ts = s["time_stamps"]
        n_samples = len(ts)
        summary = {
            "stream_name": name,
            "stream_type": stype,
            "nominal_srate": srate,
            "channel_count": channel_count,
            "channel_labels": channel_labels,
            "n_samples": n_samples,
            "timestamp_start": float(ts[0]) if n_samples else None,
            "timestamp_end": float(ts[-1]) if n_samples else None,
            "duration_sec": float(ts[-1] - ts[0]) if n_samples > 1 else None,
            "is_empty_or_malformed": n_samples == 0,
        }
        stream_summaries.append(summary)
        del s  # explicitly drop reference to the loaded time_series/time_stamps arrays
    del streams
    return {"streams": stream_summaries}


def main() -> int:
    if not DATASET_ROOT.exists():
        print(f"dataset root not found: {DATASET_ROOT}", file=sys.stderr)
        return 1

    truncated = load_truncation_set()
    print(f"loaded {len(truncated)} known-truncated paths (will be skipped for content inspection)")

    file_inventory = build_file_inventory(truncated)
    print(f"file inventory: {len(file_inventory)} files")

    with open(DOCS_DIR / "dejavu_file_inventory.csv", "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(file_inventory[0].keys()))
        w.writeheader()
        w.writerows(file_inventory)

    db_path = DATASET_ROOT / "deja_vu_database.db"
    sqlite_result = audit_sqlite(db_path) if db_path.exists() else None
    if sqlite_result:
        print(f"SQLite: {len(sqlite_result['tables'])} tables, "
              f"{sqlite_result['distinct_subject_count']} distinct subjects, "
              f"{sqlite_result['distinct_subject_session_pairs']} distinct subject-session pairs")

    xlsx_path = DATASET_ROOT / "data.xlsx"
    xlsx_result = audit_xlsx(xlsx_path) if xlsx_path.exists() else None
    if xlsx_result:
        print(f"XLSX: sheets={xlsx_result['sheet_names']}")

    channel_rows = []
    event_rows = []

    hdf5_files = [r for r in file_inventory if r["category"] in ("preprocessed_hdf5", "segment_file") and r["trusted"]]
    print(f"inspecting {len(hdf5_files)} trusted HDF5 files (metadata only)...")
    for row in hdf5_files:
        full_path = DATASET_ROOT.parent / row["relative_path"].replace("DEJA-VU/", "DEJA-VU/", 1)
        full_path = DATASET_ROOT / Path(row["relative_path"]).relative_to("DEJA-VU")
        try:
            h5info = audit_hdf5_file(full_path)
        except Exception as exc:
            print(f"  WARNING: could not inspect {row['relative_path']}: {exc}", file=sys.stderr)
            continue

        top_attrs = h5info["top_level_attrs"]
        for modality, grp in h5info["groups"].items():
            if modality == "segment_info":
                continue
            attrs = grp["attrs"]
            channel_rows.append({
                "relative_path": row["relative_path"],
                "category": row["category"],
                "modality": modality,
                "sampling_rate": attrs.get("sampling_rate", ""),
                "channels": attrs.get("channels", ""),
                "data_shape": grp["datasets"].get("data", {}).get("shape", ""),
                "timestamps_shape": grp["datasets"].get("timestamps", {}).get("shape", ""),
            })

        seg_info = h5info["groups"].get("segment_info", {}).get("attrs", {})
        if seg_info or top_attrs:
            merged = {**top_attrs, **seg_info}
            event_rows.append({
                "relative_path": row["relative_path"],
                "category": row["category"],
                "subject_attr": merged.get("subject") or merged.get("subject_id", ""),
                "session_attr": merged.get("session") or merged.get("session_id", ""),
                "type": merged.get("type", ""),
                "quadrant": merged.get("quadrant", ""),
                "video_name": merged.get("video_name", ""),
                "transition_type": merged.get("transition_type", ""),
                "start_time": merged.get("start_time", ""),
                "end_time": merged.get("end_time", ""),
                "duration": merged.get("duration", ""),
            })

    xdf_files = [r for r in file_inventory if r["category"] == "raw_xdf"]
    print(f"inspecting {len(xdf_files)} raw XDF files (streaming metadata, arrays discarded per-file)...")
    xdf_results = {}
    for row in xdf_files:
        full_path = DATASET_ROOT / Path(row["relative_path"]).relative_to("DEJA-VU")
        try:
            xdf_results[row["relative_path"]] = audit_xdf_file(full_path)
        except Exception as exc:
            print(f"  WARNING: could not parse {row['relative_path']}: {exc}", file=sys.stderr)
            xdf_results[row["relative_path"]] = {"error": str(exc)}

    with open(DOCS_DIR / "dejavu_channel_inventory.csv", "w", newline="") as f:
        if channel_rows:
            w = csv.DictWriter(f, fieldnames=list(channel_rows[0].keys()))
            w.writeheader()
            w.writerows(channel_rows)

    with open(DOCS_DIR / "dejavu_event_inventory.csv", "w", newline="") as f:
        if event_rows:
            w = csv.DictWriter(f, fieldnames=list(event_rows[0].keys()))
            w.writeheader()
            w.writerows(event_rows)

    # Subject/session summary
    subjects = sorted({r["participant"] for r in file_inventory if r["participant"]})
    sessions_by_subject: dict[str, set[str]] = {}
    for r in file_inventory:
        if r["participant"] and r["session"]:
            sessions_by_subject.setdefault(r["participant"], set()).add(r["session"])

    summary_rows = []
    for subj in subjects:
        sessions = sorted(sessions_by_subject.get(subj, []))
        raw_files = [r for r in file_inventory if r["category"] == "raw_xdf" and r["participant"] == subj]
        segment_files = [r for r in file_inventory if r["category"] == "segment_file" and r["participant"] == subj]
        preprocessed_files = [r for r in file_inventory if r["category"] == "preprocessed_hdf5" and r["participant"] == subj]
        summary_rows.append({
            "participant": subj,
            "session_count": len(sessions),
            "sessions": ";".join(sessions),
            "raw_xdf_files": len(raw_files),
            "preprocessed_hdf5_files": len(preprocessed_files),
            "segment_files": len(segment_files),
        })

    with open(DOCS_DIR / "dejavu_subject_session_summary.csv", "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(summary_rows[0].keys()))
        w.writeheader()
        w.writerows(summary_rows)

    audit_json = {
        "dataset_root": str(DATASET_ROOT),
        "file_inventory_count": len(file_inventory),
        "sqlite": sqlite_result,
        "xlsx": xlsx_result,
        "hdf5_files_inspected": len(hdf5_files),
        "hdf5_files_skipped_truncated": len([r for r in file_inventory if r["category"] in ("preprocessed_hdf5", "segment_file") and not r["trusted"]]),
        "xdf_files_inspected": len(xdf_results),
        "xdf_stream_summaries": xdf_results,
        "independent_participants": len(subjects),
        "participant_ids": subjects,
        "participant_session_pairs": sum(len(v) for v in sessions_by_subject.values()),
    }
    with open(DOCS_DIR / "data_audit_dejavu.json", "w") as f:
        json.dump(audit_json, f, indent=2, default=str)

    print("\naudit complete. wrote:")
    print("  docs/data_audit_dejavu.json")
    print("  docs/dejavu_file_inventory.csv")
    print("  docs/dejavu_subject_session_summary.csv")
    print("  docs/dejavu_channel_inventory.csv")
    print("  docs/dejavu_event_inventory.csv")
    return 0


if __name__ == "__main__":
    sys.exit(main())
